"""
Excel exporter — reads all candidate JSONs and writes one .xlsx per job role
plus a combined ALL_ROLES.xlsx (one sheet per role).

Appends new candidates to existing sheets; existing rows are never deleted.

Run standalone:  python scraper/excel_export.py
Called by:       scraper/run_scheduler.py after each scrape run
"""

import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

from scraper.utils import CANDIDATES_DIR, get_logger

log = get_logger("excel_export")

EXPORTS_DIR = CANDIDATES_DIR.parent / "exports"

# Columns written to Excel — order matters
COLUMNS = [
    ("full_name",            "Name",             40),
    ("status",               "Status",           14),
    ("job_title",            "Job Title",        30),
    ("email",                "Email",            30),
    ("phone",                "Phone",            16),
    ("geography",            "Location",         22),
    ("qualifications",       "Qualifications",   40),
    ("professional_summary", "Summary",          60),
    ("experience",           "Experience",       70),
    ("certifications",       "Certifications",   35),
    ("education",            "Education",        35),
    ("skills",               "Skills",           50),
    ("resume_text",          "Resume Text",      70),
    ("cover_letter",         "Cover Letter",     50),
    ("resume_file",          "Has Resume PDF",   16),
    ("fit_score",            "Fit Score",        12),
    ("scraped_at",           "Scraped At",       22),
    ("indeed_profile_url",   "Profile URL",      50),
    ("notes",                "Notes",            30),
]

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")
STATUS_COLORS = {
    "new":          "E2EFDA",
    "reviewing":    "FFF2CC",
    "contacting":   "FCE4D6",
    "interviewing": "DDEBF7",
    "hired":        "C6EFCE",
    "rejected":     "FFCCCC",
}


def _cell_value(record: dict, field: str) -> str:
    val = record.get(field)
    if val is None:
        return ""
    if field == "resume_file":
        return "Yes" if val else "No"
    if field == "scraped_at" and val:
        try:
            return datetime.fromisoformat(val).strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(val)
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    text = str(val).strip()
    # Truncate very long fields for readability (Excel cells can hold 32k chars)
    if field in ("experience", "resume_text") and len(text) > 5000:
        text = text[:5000] + "…"
    elif len(text) > 2000:
        text = text[:2000] + "…"
    return text


def _write_headers(ws):
    for col_idx, (field, header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(wrap_text=False, vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def _write_row(ws, row_num: int, record: dict):
    status = record.get("status", "").lower()
    row_fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
    if row_num % 2 == 0 and status not in STATUS_COLORS:
        row_fill = ALT_FILL

    for col_idx, (field, _, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=_cell_value(record, field))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if status in STATUS_COLORS:
            cell.fill = row_fill
        elif row_num % 2 == 0:
            cell.fill = ALT_FILL


def _existing_ids(ws) -> set[str]:
    """Return set of indeed IDs already in the sheet (col A = name used as proxy via URL col)."""
    # We use profile URL column as unique key (last column group)
    url_col = next(
        (i + 1 for i, (f, _, _) in enumerate(COLUMNS) if f == "indeed_profile_url"), None
    )
    ids = set()
    if url_col is None:
        return ids
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[url_col - 1]:
            ids.add(str(row[url_col - 1]))
    return ids


def export_all() -> int:
    """
    Read all candidate JSONs, group by job_title, write/update per-role xlsx
    and combined ALL_ROLES.xlsx. Returns total candidate count exported.
    """
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    json_files = sorted(CANDIDATES_DIR.glob("*.json"))
    if not json_files:
        log.warning("No candidate JSONs found in data/candidates/")
        return 0

    # Load all records
    records: list[dict] = []
    for f in json_files:
        try:
            records.append(json.loads(f.read_text(encoding="utf-8")))
        except Exception as e:
            log.warning(f"Skipping bad JSON {f.name}: {e}")

    if not records:
        log.warning("No valid records to export.")
        return 0

    # Group by job_title
    by_role: dict[str, list[dict]] = {}
    for r in records:
        role = r.get("job_title") or "Unknown"
        by_role.setdefault(role, []).append(r)

    log.info(f"Exporting {len(records)} candidates across {len(by_role)} role(s)...")

    # ── Per-role files (always rebuilt from all JSONs = old + new) ───────────
    for role, role_records in sorted(by_role.items()):
        safe_name = role.replace("/", "-").replace("\\", "-").strip()
        xlsx_path = EXPORTS_DIR / f"{safe_name}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = safe_name[:31]
        _write_headers(ws)

        current_row = 2          # row 1 = headers
        for rec in role_records:
            _write_row(ws, current_row, rec)
            current_row += 1

        wb.save(xlsx_path)
        log.info(f"  {role}: {len(role_records)} candidate(s) -> {xlsx_path.name}")

    # ── Combined ALL_ROLES.xlsx ───────────────────────────────────────────────
    all_path = EXPORTS_DIR / "ALL_ROLES.xlsx"
    wb_all   = openpyxl.Workbook()
    wb_all.remove(wb_all.active)  # remove default empty sheet

    for role, role_records in sorted(by_role.items()):
        safe = role.replace("/", "-").strip()[:31]
        ws   = wb_all.create_sheet(title=safe)
        _write_headers(ws)
        for row_num, rec in enumerate(role_records, 2):
            _write_row(ws, row_num, rec)

    wb_all.save(all_path)
    log.info(f"  Combined -> {all_path.name}  ({len(records)} rows across {len(by_role)} sheet(s))")

    return len(records)


if __name__ == "__main__":
    total = export_all()
    print(f"\nDone. {total} candidate(s) exported to data/exports/")
